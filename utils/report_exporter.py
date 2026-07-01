"""
Shared report export utility for DairyERP.
Exports any tabular data (list of dicts or sqlite3.Row list) to Excel or PDF.

Place this file at: DairyERP/utils/report_exporter.py
Create empty __init__.py in utils/ folder too.

Usage from any panel:
    from utils.report_exporter import export_to_excel, export_to_pdf

    export_to_excel(
        data=rows,                      # list of sqlite3.Row or list of dicts
        columns=["ID","Name","Amount"], # column headers
        keys=["id","name","amount"],    # matching dict/row keys (lowercase, no spaces)
        filename="Sales_Report",
        title="Sales Report"
    )
"""

import os
from datetime import datetime
from tkinter import filedialog, messagebox

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


# ==========================
# DESKTOP FOLDER HELPER
# ==========================

def get_default_save_path(filename, extension):
    """Returns path to Desktop, or home folder if Desktop not found."""
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    if not os.path.isdir(desktop):
        desktop = os.path.expanduser("~")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = filename.replace(" ", "_")
    return os.path.join(desktop, f"{safe_name}_{timestamp}.{extension}")


def _row_value(row, key, index):
    """Get value from sqlite3.Row or dict by key, fallback to index."""
    try:
        return row[key]
    except (KeyError, IndexError, TypeError):
        try:
            return row[index]
        except (KeyError, IndexError, TypeError):
            return ""


# ==========================
# EXPORT TO EXCEL
# ==========================

def export_to_excel(data, columns, keys, filename="Report",
                    title="Dairy ERP Report", parent_window=None):
    """
    data:     list of sqlite3.Row or list of dicts
    columns:  list of display column headers, e.g. ["ID","Name","Amount"]
    keys:     list of matching keys to extract from each row
    filename: base filename (without extension/timestamp)
    title:    report title shown at top of sheet

    Returns True on success, False on failure/cancel.
    """
    if not OPENPYXL_AVAILABLE:
        messagebox.showerror(
            "Missing Library",
            "openpyxl is not installed.\nRun: pip install openpyxl"
        )
        return False

    if not data:
        messagebox.showwarning("No Data", "There is no data to export.")
        return False

    default_path = get_default_save_path(filename, "xlsx")
    path = filedialog.asksaveasfilename(
        title="Save Excel Report",
        initialfile=os.path.basename(default_path),
        initialdir=os.path.dirname(default_path),
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not path:
        return False

    try:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Report"

        # Title row
        sheet.merge_cells(start_row=1, start_column=1,
                          end_row=1, end_column=len(columns))
        title_cell = sheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="1565C0")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 28

        # Generated on row
        sheet.merge_cells(start_row=2, start_column=1,
                          end_row=2, end_column=len(columns))
        gen_cell = sheet.cell(row=2, column=1,
                              value=f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
        gen_cell.font = Font(name="Arial", size=10, italic=True, color="555555")
        gen_cell.alignment = Alignment(horizontal="center")

        # Header row
        header_row = 4
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        for col_idx, col_name in enumerate(columns, start=1):
            cell = sheet.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E7D32")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Data rows
        for r_idx, row in enumerate(data, start=header_row + 1):
            for c_idx, key in enumerate(keys, start=1):
                value = _row_value(row, key, c_idx - 1)
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                cell.font = Font(name="Arial", size=10)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if r_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F5F5F5")

        # Auto column width
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = len(str(col_name))
            for row in data:
                val = str(_row_value(row, keys[col_idx-1], col_idx-1))
                max_len = max(max_len, len(val))
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        wb.save(path)
        messagebox.showinfo("✅ Exported", f"Report saved successfully!\n\n{path}")
        return True

    except Exception as e:
        messagebox.showerror("❌ Export Failed", f"Could not save Excel file.\n\n{e}")
        return False


# ==========================
# EXPORT TO PDF
# ==========================

def export_to_pdf(data, columns, keys, filename="Report",
                  title="Dairy ERP Report", subtitle="", parent_window=None):
    """
    Same signature as export_to_excel but produces a PDF.
    Auto-switches to landscape if more than 6 columns.
    """
    if not REPORTLAB_AVAILABLE:
        messagebox.showerror(
            "Missing Library",
            "reportlab is not installed.\nRun: pip install reportlab"
        )
        return False

    if not data:
        messagebox.showwarning("No Data", "There is no data to export.")
        return False

    default_path = get_default_save_path(filename, "pdf")
    path = filedialog.asksaveasfilename(
        title="Save PDF Report",
        initialfile=os.path.basename(default_path),
        initialdir=os.path.dirname(default_path),
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf")]
    )
    if not path:
        return False

    try:
        page_size = landscape(A4) if len(columns) > 6 else A4

        doc = SimpleDocTemplate(
            path, pagesize=page_size,
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=1.5*cm, bottomMargin=1.5*cm
        )

        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            "TitleStyle", parent=styles["Title"],
            fontSize=18, textColor=colors.HexColor("#1565C0"),
            spaceAfter=4
        )
        elements.append(Paragraph(f"🐄 Dairy ERP — {title}", title_style))

        if subtitle:
            sub_style = ParagraphStyle(
                "SubStyle", parent=styles["Normal"],
                fontSize=10, textColor=colors.grey
            )
            elements.append(Paragraph(subtitle, sub_style))

        gen_style = ParagraphStyle(
            "GenStyle", parent=styles["Normal"],
            fontSize=9, textColor=colors.grey, spaceAfter=12
        )
        elements.append(Paragraph(
            f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}",
            gen_style
        ))
        elements.append(Spacer(1, 8))

        # Build table data
        table_data = [columns]
        for row in data:
            table_row = [str(_row_value(row, k, i)) for i, k in enumerate(keys)]
            table_data.append(table_row)

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, 0), 9),
            ("FONTSIZE",     (0, 1), (-1, -1), 8),
            ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F5F5F5")]),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)

        doc.build(elements)
        messagebox.showinfo("✅ Exported", f"Report saved successfully!\n\n{path}")
        return True

    except Exception as e:
        messagebox.showerror("❌ Export Failed", f"Could not save PDF file.\n\n{e}")
        return False


# ==========================
# COMBINED EXPORT DIALOG (Excel + PDF buttons together)
# ==========================

def make_export_buttons(parent, get_data_func, columns, keys,
                        filename="Report", title="Report",
                        ctk_module=None):
    """
    Returns a frame with two buttons: Export Excel / Export PDF.
    get_data_func: a callable that returns the current data list when called
                   (so export always uses freshest filtered data)

    Usage in any panel:
        from utils.report_exporter import make_export_buttons
        export_frame = make_export_buttons(
            parent_frame,
            get_data_func=lambda: get_current_table_data(),
            columns=["ID","Name","Total"],
            keys=["id","name","total"],
            filename="Sales_Report",
            title="Sales Report",
            ctk_module=ctk
        )
        export_frame.pack(...)
    """
    import customtkinter as ctk

    frame = ctk.CTkFrame(parent, fg_color="transparent")

    def _excel():
        data = get_data_func()
        export_to_excel(data, columns, keys, filename, title)

    def _pdf():
        data = get_data_func()
        export_to_pdf(data, columns, keys, filename, title)

    ctk.CTkButton(
        frame, text="📊 Export Excel", width=140,
        fg_color="#1D6F42", hover_color="#14502F",
        command=_excel
    ).pack(side="left", padx=4)

    ctk.CTkButton(
        frame, text="📄 Export PDF", width=140,
        fg_color="#C62828", hover_color="#8E0000",
        command=_pdf
    ).pack(side="left", padx=4)

    return frame